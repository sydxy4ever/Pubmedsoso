from pathlib import Path
import csv

from pubmedsoso.models import Article, FreeStatus
from pubmedsoso.core.export import Exporter, COLUMNS, _article_to_row


def test_columns_tuple():
    assert isinstance(COLUMNS, tuple)
    assert len(COLUMNS) == 13
    assert COLUMNS[0] == "序号"
    assert COLUMNS[1] == "文献标题"
    assert COLUMNS[10] == "是否有免费全文"
    assert COLUMNS[11] == "是否是review"


def test_article_to_row():
    article = Article(
        id=1,
        title="Test Article",
        authors="Author A, Author B",
        journal="Nature 2024",
        doi="10.1234/test",
        pmid=12345678,
        pmcid="PMC12345",
        abstract="This is an abstract.",
        keywords="cancer; therapy",
        affiliations="University A",
        free_status=FreeStatus.FREE_PMC,
        is_review=True,
        save_path="/path/to/file.pdf",
    )
    row = _article_to_row(article, 1)
    assert row[0] == "1"
    assert row[1] == "Test Article"
    assert row[2] == "Author A, Author B"
    assert row[3] == "Nature 2024"
    assert row[4] == "10.1234/test"
    assert row[5] == "12345678"
    assert row[6] == "PMC12345"
    assert row[7] == "This is an abstract."
    assert row[8] == "cancer; therapy"
    assert row[9] == "University A"
    assert row[10] == "是"
    assert row[11] == "是"
    assert row[12] == "/path/to/file.pdf"


def test_export_xlsx_free_status_display(tmp_path):
    articles = [
        Article(title="Free PMC", free_status=FreeStatus.FREE_PMC),
        Article(title="Free Article", free_status=FreeStatus.FREE_ARTICLE),
        Article(title="Not Free", free_status=FreeStatus.NOT_FREE),
    ]
    output_path = tmp_path / "test.xlsx"
    result = Exporter.to_xlsx(articles, output_path)
    assert result.exists()
    from openpyxl import load_workbook

    wb = load_workbook(result)
    ws = wb.active
    assert ws.cell(row=2, column=11).value == "是"
    assert ws.cell(row=3, column=11).value == "否"
    assert ws.cell(row=4, column=11).value == "否"
    wb.close()


def test_export_xlsx(tmp_path):
    articles = [
        Article(
            title="Article 1",
            authors="Author 1",
            journal="Journal 2024",
            doi="10.1/1",
            pmid=111,
            pmcid="PMC111",
            abstract="Abstract 1",
            keywords="key1",
            affiliations="Affiliation 1",
            free_status=FreeStatus.FREE_PMC,
            is_review=True,
            save_path="/path/1.pdf",
        ),
        Article(
            title="Article 2",
            authors="Author 2",
            journal="Journal 2023",
            doi="10.1/2",
            pmid=222,
            pmcid="",
            abstract="Abstract 2",
            keywords="key2",
            affiliations="Affiliation 2",
            free_status=FreeStatus.NOT_FREE,
            is_review=False,
            save_path="",
        ),
    ]
    output_path = tmp_path / "output.xlsx"
    result = Exporter.to_xlsx(articles, output_path)
    assert result.exists()
    assert result == output_path
    from openpyxl import load_workbook

    wb = load_workbook(result)
    ws = wb.active
    for col_idx, col_name in enumerate(COLUMNS, 1):
        assert ws.cell(row=1, column=col_idx).value == col_name
    assert ws.cell(row=2, column=1).value == "1"
    assert ws.cell(row=2, column=2).value == "Article 1"
    assert ws.cell(row=3, column=1).value == "2"
    assert ws.cell(row=3, column=2).value == "Article 2"
    wb.close()


def test_export_csv(tmp_path):
    articles = [
        Article(
            title="Article 1",
            authors="Author 1",
            journal="Journal 2024",
            doi="10.1/1",
            pmid=111,
            pmcid="PMC111",
            abstract="Abstract 1",
            keywords="key1",
            affiliations="Affiliation 1",
            free_status=FreeStatus.FREE_PMC,
            is_review=True,
            save_path="/path/1.pdf",
        ),
        Article(
            title="Article 2",
            authors="Author 2",
            journal="Journal 2023",
            doi="10.1/2",
            pmid=222,
            pmcid="",
            abstract="Abstract 2",
            keywords="key2",
            affiliations="Affiliation 2",
            free_status=FreeStatus.NOT_FREE,
            is_review=False,
            save_path="",
        ),
    ]
    output_path = tmp_path / "output.csv"
    result = Exporter.to_csv(articles, output_path)
    assert result.exists()
    assert result == output_path
    with open(result, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    assert rows[0] == list(COLUMNS)
    assert rows[1][0] == "1"
    assert rows[1][1] == "Article 1"
    assert rows[1][10] == "是"
    assert rows[1][11] == "是"
    assert rows[2][0] == "2"
    assert rows[2][1] == "Article 2"
    assert rows[2][10] == "否"
    assert rows[2][11] == "否"


def test_export_creates_parent_directories(tmp_path):
    articles = [Article(title="Test")]
    output_path = tmp_path / "subdir" / "deep" / "output.xlsx"
    result = Exporter.to_xlsx(articles, output_path)
    assert result.exists()
    output_path_csv = tmp_path / "subdir2" / "deep" / "output.csv"
    result_csv = Exporter.to_csv(articles, output_path_csv)
    assert result_csv.exists()
